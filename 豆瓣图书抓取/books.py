import requests
import random
from lxml import etree
import os
from openpyxl import load_workbook
import time
import pandas as pd
from openpyxl.drawing.image import Image


bookinfo = {
    "ISBN": "",
    "书名": "",
    "作者": "",
    "出版社": "",
    "出版年份": "",
    "页数": "",
    "定价": "",
    "装帧": "",
    "图书封面照片地址": "",
    "豆瓣评分": "",
    "评分人数": "",
    "内容简介": "",
}

ua_list = [
            'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/535.1 (KHTML, like Gecko) Chrome/14.0.835.163 Safari/535.1',
            'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:6.0) Gecko/20100101 Firefox/6.0',
            'Mozilla/4.0 (compatible; MSIE 8.0; Windows NT 6.1; WOW64; Trident/4.0; SLCC2; .NET CLR 2.0.50727; .\
            NET CLR 3.5.30729; .NET CLR 3.0.30729; Media Center PC 6.0; .NET4.0C; InfoPath.3)', 
        ]

#临时函数 将获取的html文件保存到本地，避免重复请求
def saveHtml(isbn, html):

    filename = isbn + '.html'
    with open(filename, 'w', encoding='utf-8') as f:
        f.write(html)

#内容简介是一个数组，转换成一个字符串
def introductionToString(introductionList):

    introductionString = '\n'.join(introductionList)
    return introductionString

#下载图书封面
def downloadBookImage(url):

    os.makedirs('/home/cshi26/openai/Python_test/ISBN/image/', exist_ok=True)
    imageName = getUrlImageName(url)
    from urllib.request import urlretrieve
    urlretrieve(url, '/home/cshi26/openai/Python_test/ISBN/image/{}'.format(imageName))

#通过douban.com/isbn/isbn号查询图书的详细信息地址
def getSubject(isbn):

    filename = isbn + '.html'
    url = "http://douban.com/isbn/{}".format(isbn)
    if os.path.exists(filename):
        file = open(filename, 'r', encoding='utf-8')
        html = file.read()
        file.close()
    else:
        html = requests.get(url=url, headers={'User-Agent': random.choice(ua_list)})
        html = html.content.decode('utf-8', 'ignore')
        saveHtml(isbn, html)

    parse_html = etree.HTML(html)

    #查询的数据有
    #书名、作者、出版社、副标题、译者、出版年、页数、定价、装帧、ISBN号、图书图片、评分、评论数、内容简介

    bookImgUrl = parse_html.xpath('//*[@id="mainpic"]/a/img/@src')[0].strip()
    bookName = parse_html.xpath('//*[@id="wrapper"]/h1/span/text()')[0].strip()
    bookInfo = parse_html.xpath('//*[@id="info"]')
    bookAuthor = bookInfo[0].xpath('.//span[@class="pl" and text()=" 作者"]/following-sibling::a/text()')[0].strip()
    #译者
    # bookTranslator = bookInfo[0].xpath('.//span[@class="pl" and text()=" 译者"]/following-sibling::a/text()')[0].strip()
    #出版社
    bookPublisher = bookInfo[0].xpath('.//span[@class="pl" and text()="出版社:"]/following-sibling::a/text()')[0].strip()
    #出版年份
    bookPublisherYear = bookInfo[0].xpath('.//span[@class="pl" and text()="出版年:"]/following-sibling::text()')[0].strip()
    #页数
    bookPage = bookInfo[0].xpath('.//span[@class="pl" and text()="页数:"]/following-sibling::text()')[0].strip()
    #定价
    bookPrice = bookInfo[0].xpath('.//span[@class="pl" and text()="定价:"]/following-sibling::text()')[0].strip()
    #装帧
    bookBind = bookInfo[0].xpath('.//span[@class="pl" and text()="装帧:"]/following-sibling::text()')[0].strip()
    bookISBN = bookInfo[0].xpath('.//span[@class="pl" and text()="ISBN:"]/following-sibling::text()')[0].strip()
    #评分
    bookScore =parse_html.xpath('//*[@id="interest_sectl"]/div[1]/div[2]/strong/text()')[0].strip()
    #评分人数
    bookScorePepoleNum= parse_html.xpath('//div[@class="rating_sum"]/span/a/span/text()')[0].strip()
    #简介
    bookIntroduction = parse_html.xpath('//div[@class="intro"]/p/text()')
    bookIntroduction = introductionToString(bookIntroduction)
    # print(bookImg, bookName, bookAuthor, bookPublisher, bookPublisherYear, bookPage, bookPrice, bookBind, bookISBN, bookScore, bookScorePepoleNum, bookIntroduction)
    downloadBookImage(bookImgUrl)
    bookinfo = {
        "ISBN": bookISBN,
        "书名": bookName,
        "作者": bookAuthor,
        "出版社": bookPublisher,
        "出版年份": bookPublisherYear,
        "页数": bookPage,
        "定价": bookPrice,
        "装帧": bookBind,
        "图书封面照片地址": bookImgUrl,
        "豆瓣评分": bookScore,
        "评分人数": bookScorePepoleNum,
        "内容简介": bookIntroduction,
    }
    return bookinfo

def readExcel(filename):

    booklist = []
    workbook = load_workbook(filename=filename)
    # sheet 名称获取表格
    ws = workbook["Sheet1"]
    
    for i in range(1, ws.max_row + 1):
        # isbn =ws.cell(i, 1).value
        isbn = ws.cell(i, 1).value
        book = getSubject(isbn)
        booklist.append(book)
        # time.sleep(3)
    # for i in sheet[isbns]:
    #     isbn = i[0].value
    #     book = getSubject(isbn)
    #     booklist.append(book)
    #     # time.sleep(3)

    return booklist

def saveExcel(booklist, filename):

    pf = pd.DataFrame(list(booklist))
    filepath = pd.ExcelWriter(filename)
    pf.fillna(' ', inplace=True)
    pf.to_excel(filepath, index=False)
    filepath.save()

#处理保存后的excel 通过封面url获取图片并插入到excel中

def saveImageToExcel(filename):

    imgsize = (270 / 2, 368 / 2)
    wb = load_workbook(filename)
    ws = wb["Sheet1"]
    # img = Image('/home/cshi26/openai/Python_test/ISBN/image/{}.jpg'.format)
    for i in range(2, ws.max_row + 1):
        # isbn =ws.cell(i, 1).value
        imageUrl = ws.cell(i, 9).value
        imageName = getUrlImageName(imageUrl)
        ws.column_dimensions['M'].width = imgsize[0] * 0.14  # 修改列M的宽
        image = Image('/home/cshi26/openai/Python_test/ISBN/image/{}'.format(imageName))
        image.width, image.height = imgsize
        imageData = ws.cell(row=i, column=13)

        ws.add_image(image, imageData.coordinate)
        ws.row_dimensions[i].height = imgsize[1] * 0.78
        wb.save(filename)
        print(imageName, imageUrl)

#通过imageurl 获取图片名称
def getUrlImageName(url):

    return url.split('/')[-1]

if __name__ == '__main__':

    # getSubject('9787536097261')
    # try:
    #     while True:
    #         isbn = input()
    #         getSubject(isbn)
    # except:
    #     pass
    readFilename = "/home/cshi26/openai/Python_test/ISBN/Book1.xlsx"
    saveFilename = "/home/cshi26/openai/Python_test/ISBN/Book1-1.xlsx"
    booklist = readExcel(readFilename)
    saveExcel(booklist, saveFilename)
    saveImageToExcel(saveFilename)
